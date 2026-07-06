"""Ingestion: read the source Excel file and write normalized records to SQLite.

Runnable in isolation (does not embed, does not start the MCP server):

    uv run consult-ingest
    uv run python -m consult_agent.ingest --xlsx path/to/file.xlsx --db consult.db

Data-quality issues are reported to stderr rather than silently dropped or
guessed at (per CLAUDE.md working conventions).
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import openpyxl

from . import config, db
from .normalize import normalize_row

SOURCE_SHEET = "Research Data Services Consult "  # first sheet (verified)


def _read_rows(xlsx_path: Path) -> list[dict[str, object]]:
    """Read the source sheet into a list of header->value dicts."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[SOURCE_SHEET] if SOURCE_SHEET in wb.sheetnames else wb.worksheets[0]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    rows = [dict(zip(header, values)) for values in rows_iter]
    wb.close()
    return rows


def _insert(conn, record) -> None:
    conn.execute(
        """
        INSERT INTO consults (
            id, date, status, affiliation, role, brief_description,
            notes_combined, topics, assigned_to, initial_response_by,
            initial_response_date, is_incomplete, embed_text,
            netid, first_name, last_name, email, work_for
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            record.id,
            record.date,
            record.status,
            record.affiliation,
            record.role,
            record.brief_description,
            record.notes_combined,
            json.dumps(record.topics),
            record.assigned_to,
            record.initial_response_by,
            record.initial_response_date,
            1 if record.is_incomplete else 0,
            record.embed_text,
            record.netid,
            record.first_name,
            record.last_name,
            record.email,
            record.work_for,
        ),
    )


def ingest(xlsx_path: Path, db_path: Path) -> None:
    rows = _read_rows(xlsx_path)
    print(f"Read {len(rows)} data rows from {xlsx_path.name}", file=sys.stderr)

    conn = db.connect(db_path)
    db.init_schema(conn)
    # Fresh rebuild: this script owns the store and is meant to be re-runnable.
    conn.execute("DELETE FROM consults")
    conn.execute("DELETE FROM consult_vec")

    seen_ids: set[str] = set()
    inserted = 0
    skipped = 0
    incomplete = 0
    issue_rows = 0

    for i, raw in enumerate(rows, start=2):  # +2: header row + 1-based
        result = normalize_row(raw)
        rec = result.record

        if not rec.id:
            print(f"  row {i}: SKIPPED — {'; '.join(result.issues)}", file=sys.stderr)
            skipped += 1
            continue
        if rec.id in seen_ids:
            print(f"  row {i}: SKIPPED — duplicate id {rec.id!r}", file=sys.stderr)
            skipped += 1
            continue

        if result.issues:
            issue_rows += 1
            print(f"  row {i} (id={rec.id}): {'; '.join(result.issues)}", file=sys.stderr)

        seen_ids.add(rec.id)
        _insert(conn, rec)
        inserted += 1
        if rec.is_incomplete:
            incomplete += 1

    conn.commit()
    conn.close()

    print(
        f"\nIngest complete: {inserted} inserted, {skipped} skipped, "
        f"{issue_rows} rows with data-quality notes, "
        f"{incomplete} flagged is_incomplete. DB: {db_path}",
        file=sys.stderr,
    )


def main() -> None:
    p = argparse.ArgumentParser(description="Ingest consult Excel file into SQLite.")
    p.add_argument("--xlsx", type=Path, default=config.SOURCE_XLSX, help="source .xlsx")
    p.add_argument("--db", type=Path, default=config.DB_PATH, help="output SQLite file")
    args = p.parse_args()
    ingest(args.xlsx, args.db)


if __name__ == "__main__":
    main()
